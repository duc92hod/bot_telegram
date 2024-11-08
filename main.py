from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, CallbackContext
from telegram.ext import filters
from openpyxl import Workbook, load_workbook
import os
import asyncio
import requests
from bs4 import BeautifulSoup
import random

# Đặt token của bạn ở đây
YOUR_BOT_TOKEN = os.getenv("YOUR_BOT_TOKEN")  # Thay thế bằng token thực tế của bạn

# Thiết lập proxy
# Thiết lập proxy
# Lấy danh sách proxy từ biến môi trường
proxy_list_string = os.getenv("PROXY_LIST", "")
proxies = proxy_list_string.split(',')  # Tách chuỗi thành danh sách dựa trên dấu phẩy
proxies_list = [
    {
        'http': proxy.strip(),
        'https': proxy.strip(),
    }
    for proxy in proxies if proxy.strip()  # Loại bỏ proxy trống
]

# Tạo file Excel nếu không tồn tại
excel_file = "messages.xlsx"
if not os.path.exists(excel_file):
    wb = Workbook()
    wb.save(excel_file)

user_status = {}
user_messages = {}
user_exported_index = {}
user_timers = {}
user_titles = {}
user_file_status = {}  # Trạng thái file đã gửi

async def export_user_messages(update, context, user_id):
    user_file = f"{user_id}_messages.xlsx"
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = str(user_id)
    new_ws.append(["User ID", "Message", "Title",
                    "Image 1", "Image 2", "Image 3",
                    "Image 4", "Image 5", "Image 6",
                    "Image 7", "Image 8", "Image 9"])  # Thêm tiêu đề cho 9 cột hình ảnh

    if user_id in user_messages:
        for idx, msg in enumerate(user_messages[user_id][user_exported_index[user_id]:]):
            title = clean_string(user_titles[user_id][idx][0] if idx < len(user_titles[user_id]) else "Không có tiêu đề")
            img_urls = user_titles[user_id][idx][1] if idx < len(user_titles[user_id]) else []

            # Ghi vào file Excel, lưu tất cả các URL hình ảnh
            img_row = img_urls + [""] * (12 - len(img_urls))  # Thêm ô trống nếu có ít hơn 9 ảnh
            new_ws.append([user_id, msg, title] + img_row)  # Ghi vào Excel

    new_wb.save(user_file)

    with open(user_file, 'rb') as file:
        await context.bot.send_document(chat_id=update.message.chat.id, document=file)

async def stop_bot(update: Update, context: CallbackContext, user_id):
    await asyncio.sleep(1800)  # Thời gian chờ 30 phút
    if user_id in user_status and user_status[user_id]:
        if user_exported_index[user_id] < len(user_messages[user_id]):
            await export_user_messages(update, context, user_id)
        await update.message.reply_text("Bot sẽ dừng do không hoạt động trong 30 phút.")
        user_status[user_id] = False
        user_messages[user_id] = []
        user_titles[user_id] = []
        user_file_status[user_id] = False  # Đặt lại trạng thái file

async def start(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    user_status[user_id] = True
    user_messages[user_id] = []
    user_titles[user_id] = []  # Khởi tạo danh sách tiêu đề và ảnh
    user_exported_index[user_id] = 0
    user_file_status[user_id] = False  # Khởi tạo trạng thái file
    await update.message.reply_text('Xin chào! Tôi sẽ lưu tin nhắn của bạn vào file Excel từ bây giờ. Bắt đầu gửi tin nhắn!')

    if user_id in user_timers:
        user_timers[user_id].cancel()
    user_timers[user_id] = asyncio.create_task(stop_bot(update, context, user_id))

async def thongtin(update: Update, context: CallbackContext) -> None:
    commands_info = (
        "/start - Bắt đầu lưu tin nhắn.",
        "/export - Xuất tin nhắn đã lưu.",
        "/thongtin - Hiển thị thông tin về các lệnh hỗ trợ."
        #"/readfile - Đọc file Excel và lấy dữ liệu sản phẩm."
    )
    await update.message.reply_text("\n".join(commands_info))

async def export(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    await export_user_messages(update, context, user_id)
    user_exported_index[user_id] = len(user_messages[user_id])

async def fetch_url_data(url):
    max_retries = 3  # Số lần thử lại nếu không thành công
    for _ in range(max_retries):
        proxy = random.choice(proxies_list)
        try:
            response = requests.get(url, proxies=proxy, timeout=30)
            response.raise_for_status()  # Kiểm tra mã trạng thái

            soup = BeautifulSoup(response.content, 'html.parser')

            # Lấy tiêu đề từ thẻ <div> có class 'index-title--AnTxK'
            title_div = soup.find('div', class_='index-title--AnTxK')
            title = title_div.get_text() if title_div else "Không tìm thấy tiêu đề."

            # Kiểm tra sự tồn tại của slick-track và lấy hình ảnh
            slick_track = soup.find('div', class_='slick-track')
            img_tags = slick_track.find_all('img') if slick_track else []
            img_urls = [img.get('src') for img in img_tags if img.get('src')]
            if not img_urls:
                img_urls = ["Không tìm thấy hình ảnh."]

            return title, img_urls

        except requests.exceptions.RequestException as e:
            print(f"Không thể truy cập trang: {e}")
            continue
    return "Không lấy được dữ liệu", []

async def echo(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    message_text = update.message.text

    if user_status.get(user_id, False):
        user_messages[user_id].append(message_text)

        # Kiểm tra nếu tin nhắn là một URL
        if message_text.startswith('http://') or message_text.startswith('https://'):
            title, img_urls = await fetch_url_data(message_text)
            user_titles[user_id].append((title, img_urls))
            num_images = len(img_urls)  # Đếm số lượng hình ảnh
            
            # Gửi thông báo lưu tin nhắn
            await update.message.reply_text(f'Tin nhắn của bạn đã được lưu: "{message_text}"\nTiêu đề: {title}\nSố lượng hình ảnh: {num_images}')
        else:
            user_titles[user_id].append(("Không phải link", []))
            await update.message.reply_text(f'Tin nhắn của bạn đã được lưu: "{message_text}"')

        # Thiết lập lại thời gian chờ
        if user_id in user_timers and user_timers[user_id]:
            user_timers[user_id].cancel()
        user_timers[user_id] = asyncio.create_task(stop_bot(update, context, user_id))
    else:
        await update.message.reply_text("Vui lòng sử dụng lệnh /start để bắt đầu lưu tin nhắn.")

async def read_excel_file(file_path):
    """Đọc file Excel và lấy các link sản phẩm từ cột 'Link'."""
    df = load_workbook(file_path)
    results = []

    for sheet in df.sheetnames:
        worksheet = df[sheet]
        for row in worksheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua hàng tiêu đề
            url = row[0]  # Giả sử link nằm ở cột đầu tiên
            if not url or url.strip() == "":
                continue
            await asyncio.sleep(random.uniform(1, 2))
            title, img_urls = await fetch_url_data(url)
            results.append({'URL': url, 'Title': title, 'Images': img_urls})

    return results
def clean_string(value):
    if isinstance(value, str):
        # Loại bỏ ký tự không hợp lệ
        return ''.join(char for char in value if char.isprintable())
    return value
MAX_FILE_SIZE = 5 * 1024 * 1024
async def read_file(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id

    # Kiểm tra trạng thái người dùng đã bắt đầu tương tác
    if not user_status.get(user_id, False):
        await update.message.reply_text("Vui lòng sử dụng lệnh /start để bắt đầu.")
        return

    # Kiểm tra nếu có file được gửi
    if update.message.document:
        if update.message.document.file_size > MAX_FILE_SIZE:
            await update.message.reply_text("File quá lớn, vui lòng gửi file nhỏ hơn 5MB.")
            return
        
        file = await update.message.document.get_file()
        input_file = f"{user_id}_input_file.xlsx"
        try:
            await file.download_to_drive(input_file)  # Tải file về
            await update.message.reply_text("Hãy chờ tôi xử lý file của bạn...")
            results = await read_excel_file(input_file)
        except Exception as e:
            await update.message.reply_text(f"Lỗi khi đọc file: {e}")
            return
        # Tạo file Excel mới với kết quả
        output_file = f"{user_id}_output.xlsx"
        output_wb = Workbook()
        output_ws = output_wb.active
        output_ws.append(["URL", "Title"] + [f"Image URL {i+1}" for i in range(max(len(result['Images']) for result in results))])  # Tiêu đề cột
        count = 0  # Biến đếm số lượng liên kết đã xử lý
        try:
            for result in results:
                row = [result['URL'], clean_string(result['Title'])]
                row.extend(result['Images'])  # Thêm từng link hình ảnh vào hàng
                output_ws.append(row)
                count += 1  # Tăng biến đếm
                if count % 20 == 0:
                    await update.message.reply_text(f"Đã xử lý {count} liên kết...")
                    await asyncio.sleep(0.5)
                # Lưu file tạm sau mỗi liên kết được xử lý
                output_wb.save(output_file)

        except Exception as e:
            await update.message.reply_text(f"Đã xảy ra lỗi sau khi xử lý {count} liên kết: {e}")
            # Sau khi có lỗi, gửi file với những gì đã xử lý
            with open(output_file, 'rb') as f:
                await context.bot.send_document(chat_id=update.message.chat.id, document=f)
            return
        # Sau khi xử lý hết, gửi file kết quả hoàn chỉnh
        output_wb.save(output_file)

        with open(output_file, 'rb') as f:
            await context.bot.send_document(chat_id=update.message.chat.id, document=f)
        # Đánh dấu là đã gửi file
        user_file_status[user_id] = True  # Đánh dấu là đã gửi file
        
        await update.message.reply_text("File đã được xử lý. Bạn có thể gửi file mới bất kỳ lúc nào.")
    else:
        await update.message.reply_text("Vui lòng gửi file Excel để xử lý.")

async def export(update: Update, context: CallbackContext) -> None:
    user_id = update.message.from_user.id
    await export_user_messages(update, context, user_id)
    user_exported_index[user_id] = len(user_messages[user_id])
def main() -> None:
    application = ApplicationBuilder().token(YOUR_BOT_TOKEN).build()

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("thongtin", thongtin))
    application.add_handler(MessageHandler(filters.Document.ALL & ~filters.COMMAND, read_file))  # Xử lý file tải lên
    application.add_handler(CommandHandler("export", export))
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, echo))

    application.run_polling()

if __name__ == '__main__':
    main()